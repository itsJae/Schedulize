from datetime import date

import pytest
from openpyxl import load_workbook

from schedulize.duty_config import DEFAULT_DUTIES
from schedulize.excel_writer import write_workbook
from schedulize.models import Rank, Soldier
from schedulize.scheduler import build_schedule

DUTY_NAMES = [d.name for d in DEFAULT_DUTIES]


@pytest.fixture
def workbook_path(tmp_path):
    roster = [
        Soldier(f"24-{i:04d}", f"Soldier {i}", Rank.CORPORAL, date(2030, 1, 1), "")
        for i in range(20)
    ]
    roster.append(
        Soldier("24-9999", "Pvt Park", Rank.PRIVATE, date(2030, 1, 1), "")
    )
    schedule = build_schedule(roster, 2026, 9, DEFAULT_DUTIES)
    path = tmp_path / "duty.xlsx"
    write_workbook(schedule, path)
    return path


def test_one_sheet_per_duty_plus_summary_sheets(workbook_path):
    wb = load_workbook(workbook_path)
    assert wb.sheetnames == DUTY_NAMES + ["Duty Count", "Roster"]


def test_each_duty_sheet_has_its_own_shift_columns(workbook_path):
    wb = load_workbook(workbook_path)
    for duty in DEFAULT_DUTIES:
        header = [c.value for c in wb[duty.name][1]]
        assert header == ["Date", "Day"] + list(duty.shifts)


def test_duty_sheet_rows_cover_month(workbook_path):
    ws = load_workbook(workbook_path)["Night Watch"]
    assert ws.max_row == 31  # header + 30 September days
    first_data = [c.value for c in ws[2]]
    assert first_data[0] == "2026-09-01"
    assert first_data[1] == "Tue"
    assert "CPL" in first_data[2]


def test_holiday_and_weekend_rows_labeled_on_every_duty_sheet(workbook_path):
    wb = load_workbook(workbook_path)
    for name in DUTY_NAMES:
        rows = {row[0].value: row for row in wb[name].iter_rows(min_row=2)}
        assert "Chuseok" in str(rows["2026-09-25"][2].value)
        assert "Weekend" in str(rows["2026-09-05"][2].value)


def test_duty_count_sheet_has_per_duty_and_total_columns(workbook_path):
    ws = load_workbook(workbook_path)["Duty Count"]
    assert [c.value for c in ws[1]] == (
        ["ID", "Rank", "Name"] + DUTY_NAMES + ["Total"]
    )
    assert ws.max_row == 22  # header + 21 soldiers
    rows = {row[0].value: [c.value for c in row] for row in ws.iter_rows(min_row=2)}
    private = rows["24-9999"]
    assert private[3:] == [0, 0, 0, 0, 0]
    a_soldier = rows["24-0000"]
    assert a_soldier[-1] == sum(a_soldier[3:-1])
    assert a_soldier[-1] > 0


def test_roster_sheet_shows_exclusions(workbook_path):
    ws = load_workbook(workbook_path)["Roster"]
    rows = {row[0].value: [c.value for c in row] for row in ws.iter_rows(min_row=2)}
    assert rows["24-9999"][4] == "Excluded"
    assert "rank" in rows["24-9999"][5]
    assert rows["24-0000"][4] == "Eligible"
