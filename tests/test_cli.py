import pytest
from openpyxl import load_workbook

from schedulize.cli import main

CSV = (
    "id,name,rank,retirement_date,medical_condition\n"
    + "".join(
        f"24-{i:04d},Soldier {i},CPL,2030-01-01,\n" for i in range(14)
    )
)

DEFAULT_SHEETS = ["Night Watch", "Guard Duty", "CCTV", "Dishwashing",
                  "Duty Count", "Roster"]


@pytest.fixture
def roster_file(tmp_path):
    path = tmp_path / "roster.csv"
    path.write_text(CSV, encoding="utf-8")
    return path


def test_generates_workbook_with_default_duties(roster_file, tmp_path):
    out = tmp_path / "out.xlsx"
    assert main([str(roster_file), "--month", "2026-09", "-o", str(out)]) == 0
    assert load_workbook(out).sheetnames == DEFAULT_SHEETS


def test_duties_option_overrides_defaults(roster_file, tmp_path):
    duties = tmp_path / "duties.json"
    duties.write_text(
        '[{"name": "CCTV", "shifts": ["00-06"], "soldiers_per_shift": 1}]',
        encoding="utf-8",
    )
    out = tmp_path / "out.xlsx"
    assert main(
        [str(roster_file), "--month", "2026-09", "-o", str(out),
         "--duties", str(duties)]
    ) == 0
    assert load_workbook(out).sheetnames == ["CCTV", "Duty Count", "Roster"]


def test_duties_json_in_cwd_is_used_automatically(roster_file, tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    (tmp_path / "duties.json").write_text(
        '[{"name": "Gate Guard", "shifts": ["06-18"], "soldiers_per_shift": 2}]',
        encoding="utf-8",
    )
    out = tmp_path / "out.xlsx"
    assert main([str(roster_file), "--month", "2026-09", "-o", str(out)]) == 0
    assert load_workbook(out).sheetnames == ["Gate Guard", "Duty Count", "Roster"]


def test_bad_duties_file_reported(roster_file, tmp_path, capsys):
    duties = tmp_path / "duties.json"
    duties.write_text('[{"name": "CCTV"}]', encoding="utf-8")
    assert main(
        [str(roster_file), "--month", "2026-09", "--duties", str(duties)]
    ) == 1
    assert "shifts" in capsys.readouterr().err


def test_default_output_name(roster_file, tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    assert main([str(roster_file), "--month", "2026-09"]) == 0
    assert (tmp_path / "duty_schedule_2026-09.xlsx").exists()


def test_bad_month_format_fails(roster_file, capsys):
    assert main([str(roster_file), "--month", "September"]) == 1
    assert "YYYY-MM" in capsys.readouterr().err


def test_missing_file_fails(tmp_path, capsys):
    missing = tmp_path / "nope.csv"
    assert main([str(missing), "--month", "2026-09"]) == 1
    assert "not found" in capsys.readouterr().err


def test_roster_errors_reported(tmp_path, capsys):
    path = tmp_path / "bad.csv"
    path.write_text(
        "id,name,rank,retirement_date,medical_condition\n"
        "24-1,Kim,GENERAL,2030-01-01,\n",
        encoding="utf-8",
    )
    assert main([str(path), "--month", "2026-09"]) == 1
    assert "GENERAL" in capsys.readouterr().err


def test_infeasible_schedule_reported(tmp_path, capsys):
    path = tmp_path / "few.csv"
    path.write_text(
        "id,name,rank,retirement_date,medical_condition\n"
        "24-1,Kim,CPL,2030-01-01,\n",
        encoding="utf-8",
    )
    assert main([str(path), "--month", "2026-09"]) == 1
    assert "eligible" in capsys.readouterr().err
