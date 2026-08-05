"""Render a MonthSchedule to a styled Excel workbook, one sheet per duty."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from schedulize.duty_config import DutyType
from schedulize.scheduler import MonthSchedule

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WEEKEND_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
HOLIDAY_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
CENTER = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(*(Side(style="thin", color="BFBFBF"),) * 4)

# Excel sheet names: max 31 chars, no []:*?/\
FORBIDDEN_SHEET_CHARS = str.maketrans("", "", "[]:*?/\\")


def _sheet_title(name: str) -> str:
    return name.translate(FORBIDDEN_SHEET_CHARS)[:31]


def write_workbook(schedule: MonthSchedule, path: str | Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for duty in schedule.duties:
        _write_duty_sheet(wb.create_sheet(_sheet_title(duty.name)), schedule, duty)
    _write_count_sheet(wb.create_sheet("Duty Count"), schedule)
    _write_roster_sheet(wb.create_sheet("Roster"), schedule)
    wb.save(path)


def _format_names(soldiers) -> str:
    return " / ".join(f"{s.rank.value} {s.name}" for s in soldiers)


def _write_duty_sheet(ws, schedule: MonthSchedule, duty: DutyType) -> None:
    ws.append(["Date", "Day"] + list(duty.shifts))

    first_shift_col = 3
    last_shift_col = 2 + len(duty.shifts)

    for day in schedule.days:
        row_index = ws.max_row + 1
        row = [day.date.strftime("%Y-%m-%d"), day.date.strftime("%a")]

        if day.status == "duty":
            shifts = day.assignments[duty.name]
            row += [_format_names(shifts[shift]) for shift in duty.shifts]
            fill = None
        else:
            label = "Weekend" if day.status == "weekend" else "Holiday"
            if day.holiday_name:
                label += f" — {day.holiday_name}"
            row += [label] + [""] * (len(duty.shifts) - 1)
            fill = WEEKEND_FILL if day.status == "weekend" else HOLIDAY_FILL

        ws.append(row)
        if fill:
            if last_shift_col > first_shift_col:
                ws.merge_cells(
                    start_row=row_index,
                    start_column=first_shift_col,
                    end_row=row_index,
                    end_column=last_shift_col,
                )
            for cell in ws[row_index]:
                cell.fill = fill

    _style_sheet(ws, widths={"A": 12, "B": 6})


def _write_count_sheet(ws, schedule: MonthSchedule) -> None:
    duty_names = [d.name for d in schedule.duties]
    ws.append(["ID", "Rank", "Name"] + duty_names + ["Total"])
    ordered = sorted(
        schedule.soldiers,
        key=lambda s: (-schedule.total_counts[s.id], s.name),
    )
    for soldier in ordered:
        per_duty = [
            schedule.per_duty_counts[name][soldier.id] for name in duty_names
        ]
        ws.append(
            [soldier.id, soldier.rank.value, soldier.name]
            + per_duty
            + [schedule.total_counts[soldier.id]]
        )
    _style_sheet(ws)


def _write_roster_sheet(ws, schedule: MonthSchedule) -> None:
    ws.append(["ID", "Rank", "Name", "Retirement Date", "Status", "Reason"])
    month_start = schedule.days[0].date
    for soldier in schedule.soldiers:
        reason = soldier.exclusion_reason(month_start)
        ws.append(
            [
                soldier.id,
                soldier.rank.value,
                soldier.name,
                soldier.retirement_date.isoformat(),
                "Excluded" if reason else "Eligible",
                reason or "",
            ]
        )
    _style_sheet(ws)


def _style_sheet(ws, widths: dict[str, int] | None = None) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    for column_index in range(1, ws.max_column + 1):
        letter = get_column_letter(column_index)
        if widths and letter in widths:
            ws.column_dimensions[letter].width = widths[letter]
            continue
        longest = max(
            (len(str(c.value)) for c in ws[letter] if c.value is not None),
            default=8,
        )
        ws.column_dimensions[letter].width = min(longest + 4, 40)

    ws.freeze_panes = "A2"
